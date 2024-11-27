import os
# import params
import random
import simpy
import datetime
import copy
import  time
import numpy as np
import argparse
import openpyxl
from PPO.rewards import *
import matplotlib.pyplot as plt
from Environment.hy_jobSop import JobShop
from newMethods.HDRLMS import HiPPOS
def main(algorithms, args):

    num_test_instances = 20
    all_results = np.zeros((num_test_instances, 4, len(algorithms)))
    results = np.zeros((4, len(algorithms)))
    WTmean = np.zeros((num_test_instances, len(algorithms)))
    WTmax = np.zeros((num_test_instances, len(algorithms)))
    WFmean = np.zeros((num_test_instances, len(algorithms)))
    machine_UR = np.zeros((num_test_instances, len(algorithms)))
    test_time = np.zeros((num_test_instances, len(algorithms)))
    for num in range(num_test_instances):
        seed = random.randint(0, 1000000)
        # seed = 12
        results = np.zeros((4, len(algorithms)))
        env = simpy.Environment()
        avg = np.average(args.processing_time_range) - 0.5
        beta = avg / args.E_utliz
        np.random.seed(seed)
        arrival_interval = np.random.exponential(beta, args.num_new_jobs).round() + 1
        sim = JobShop(env, args, arrival_interval, seed)
        sim.starTime = np.random.randint(1, 2000)
        sim.decision_points.append(0)
        sim.decision_points.append(arrival_interval[0])
        sim.decision_points.append(sim.starTime)
        sim.decision_points = sorted(sim.decision_points)

        for index, algorithm in enumerate(algorithms):
            test_sim = copy.deepcopy(sim)
            start_time = time.time()
            objectives = eval(algorithm)(test_sim, args)
            end_time = time.time()
            test_time[num, index] = end_time - start_time
            WTmean[num, index] = objectives[0]
            WTmax[num, index] = objectives[1]
            WFmean[num, index] = objectives[2]
            machine_UR[num, index] = objectives[3]

    results[0] = np.mean(WTmean, axis=0)
    results[1] = np.mean(WTmax, axis=0)
    results[2] = np.mean(WFmean, axis=0)
    results[3] = np.mean(machine_UR, axis=0)
    return results, [WTmean, WTmax, WFmean, machine_UR], test_time

if __name__ == '__main__':
    parser = argparse.ArgumentParser("Hyperparameter Setting for PPO-discrete")
    parser.add_argument("--max_train_steps", type=int, default=int(1e3),
                        help=" Maximum number of training steps")  # 2e5
    parser.add_argument("--evaluate_freq", type=float, default=5e3,
                        help="Evaluate the policy every 'evaluate_freq' steps")
    parser.add_argument("--save_freq", type=int, default=20, help="Save frequency")
    parser.add_argument("--num_test_instances", type=int, default=20, help="the number of test instances")
    parser.add_argument("--batch_size", type=int, default=32, help="Batch size")
    parser.add_argument("--global_batch_size", type=int, default=32, help="Batch size")
    parser.add_argument("--mini_batch_size", type=int, default=48, help="Minibatch size")
    parser.add_argument("--state_dim", type=int, default=12, help="The dimension of local state")
    parser.add_argument("--task_state_dim", type=int, default=5, help="The dimension of task")
    parser.add_argument("--num_input", type=int, default=5, help="The types of input")
    parser.add_argument("--machine_state_dim", type=int, default=5, help="The dimension of machine")
    parser.add_argument("--action_state_dim", type=int, default=3, help="The dimension of action")
    parser.add_argument("--global_state_dim", type=int, default=24, help="The dimension of global state")
    parser.add_argument("--machine_embedding_dim", type=int, default=5, help="The embedding dimension of machine")
    parser.add_argument("--RP_dim", type=int, default=5, help="The dimension of replay buffer")
    parser.add_argument("--jobs_embedding_dim", type=int, default=5, help="The embedding dimension of jobs")
    parser.add_argument("--num_hidden_layer", type=int, default=2, help="The number of hidden layer")  # 3
    parser.add_argument("--embedding_dim", type=int, default=20, help="The number of hidden layer")  # 50
    parser.add_argument("--hidden_width", type=int, default=10,
                        help="The number of neurons in hidden layers of the neural network")
    parser.add_argument("--hidden_layers", type=int, default=5,
                        help="The number of hidden layers in the actor policy network")
    parser.add_argument("--task_hidden_width", type=int, default=30,
                        help="The number of neurons in hidden layers of the neural network")
    parser.add_argument("--actor_hidden_layers", type=int, default=5,
                        help="The number of hidden layers in the actor policy network")
    parser.add_argument("--critic_hidden_layers", type=int, default=3,
                        help="The number of hidden layers in the critic policy network")
    parser.add_argument("--action_embedding_width", type=int, default=1,
                        help="The number of neurons in hidden layers of the neural network")
    parser.add_argument("--lr_a", type=float, default=6e-3, help="Learning rate of actor")
    parser.add_argument("--lr_c", type=float, default=1e-3, help="Learning rate of critic")
    parser.add_argument("--control_lr_a", type=float, default=1e-5, help="Learning rate of control_actor")
    parser.add_argument("--control_lr_c", type=float, default=1e-4, help="Learning rate of control_critic")
    parser.add_argument("--gamma", type=float, default=0.99, help="Discount factor")
    parser.add_argument("--lamda", type=float, default=0.95, help="GAE parameter")
    parser.add_argument("--epsilon", type=float, default=0.2, help="PPO clip parameter")
    parser.add_argument("--K_epochs", type=int, default=10, help="PPO parameter")

    parser.add_argument("--num_tasks", type=int, default=3, help="The number of tasks")
    parser.add_argument("--num_total_jobs", type=int, default=100, help="The number of jobs in system")
    parser.add_argument("--num_new_jobs", type=int, default=20, help="The number of new jobs")
    parser.add_argument("--num_machines", type=int, default=10, help="The number of machines")
    parser.add_argument("--n_action", type=int, default=8, help="The number of dispatching rules")
    parser.add_argument("--E_utliz", type=float, default=0.95, help="The machine utilization ")
    parser.add_argument("--num_ops_range", type=tuple, default=(1, 10), help="The range of operations in a job")
    parser.add_argument("--num_cand_machines_range", type=tuple, default=(1, 10),
                        help="The range of candidate machines for each operation")
    parser.add_argument("--weights", type=list, default=[0.2, 0.6, 0.2], help="The weight of each job")
    parser.add_argument("--processing_time_range", type=tuple, default=(1, 99),
                        help="The processing time of an operation")
    parser.add_argument("--due_time_multiplier", type=float, default=1.5, help="The due time multiplier of a job")
    parser.add_argument("--num_warmup_jobs", type=int, default=10, help="The number of warmup jobs")
    parser.add_argument("--seed", type=int, default=12, help="seed")
    parser.add_argument("--C", type=int, default=5, help="the update frequencey of global policy")

    parser.add_argument("--use_adv_norm", type=bool, default=True, help="Trick 1:advantage normalization")
    parser.add_argument("--use_state_norm", type=bool, default=False, help="Trick 2:state normalization")
    parser.add_argument("--use_reward_norm", type=bool, default=False, help="Trick 3:reward normalization")
    parser.add_argument("--use_reward_scaling", type=bool, default=True, help="Trick 4:reward scaling")
    parser.add_argument("--entropy_coef", type=float, default=0.01, help="Trick 5: policy entropy")
    parser.add_argument("--use_lr_decay", type=bool, default=True, help="Trick 6:learning rate Decay")
    parser.add_argument("--use_grad_clip", type=bool, default=True, help="Trick 7: Gradient clip")
    parser.add_argument("--use_orthogonal_init", type=bool, default=True, help="Trick 8: orthogonal initialization")
    parser.add_argument("--set_adam_eps", type=float, default=True, help="Trick 9: set Adam epsilon=1e-5")
    parser.add_argument("--use_tanh", type=float, default=False, help="Trick 10: tanh activation function")
    parser.add_argument("--use_attention_fusion", type=float, default=True,
                        help="attention fusion for controller agent")

    parser.add_argument("--testModels_dir", type=str, default='E:/Phd_work/myWork/Code/2/testModels/',
                        help="Save path of the model")
    args = parser.parse_args()
    algorithms = ['HiPPOS']
    results, objectives, test_time = main(algorithms, args)

    print(results)
    figure_file = 'E:/Phd_work/myWork/Code/2/results/figures/'
    try:
        os.mkdir(figure_file)
        print(f"Folder '{figure_file}' created successfully.")
    except OSError as error:
        print(f"Creation of folder '{figure_file}' failed. {error}")
    y_labels = ['WT_mean', 'WT_max', 'WF_mean', 'machine_UR']
    colors = ['blue', 'orange', 'green', 'red']

    fig, axs = plt.subplots(2, 2, figsize=(12, 8))
    count = 0
    X = range(len(algorithms))
    for i in range(2):
        for j in range(2):
            axs[i, j].set_xlabel('algorithms')
            axs[i, j].set_ylabel(y_labels[count])
            axs[i, j].set_title(f"{y_labels[count]} of different algorithms")
            axs[i, j].plot(X, results[count])
            axs[i, j].set_xticks(X, labels=algorithms, rotation=45)
            count += 1
    plt.tight_layout()
    now = datetime.datetime.now()
    now_time = str(now.month) + '_' + str(now.day) + '_' + str(now.hour) + '_' + str(now.minute)
    figure_file = figure_file + now_time + '_' + str(args.num_new_jobs) + '_results.png'
    plt.savefig(figure_file)

    training_data_recoder = 'E:/Phd_work/myWork/Code/2/results/tables/'
    # now_time = datetime.datetime.now()
    scale = str(args.num_new_jobs) + '_' + str(args.num_machines)
    saved_path = training_data_recoder + scale
    try:
        os.mkdir(saved_path)
        print(f"Folder '{saved_path}' created successfully.")
    except OSError as error:
        print(f"Creation of folder '{saved_path}' failed. {error}")
    now_file_name = saved_path + '/' + now_time + '_training_objectives.xlsx'
    # xls = xlwt.Workbook()
    xls = openpyxl.Workbook()
    for ii in range(args.num_tasks+1):
        sht = xls.create_sheet()
        if ii < 3:
            sht.title = 'task' + str(ii)
        if ii == 3:
            sht.title = 'MUR'
        sht.cell(1, 1, y_labels[ii])
        for c in range(1, args.num_test_instances + 1):
            sht.cell(1, c + 1, 'instance' + str(c))
        sht.cell(1, c + 2, 'max')
        sht.cell(1, c + 3, 'min')
        sht.cell(1, c + 4, 'average')
        sht.cell(1, c + 5, 'std')
        sht.cell(1, c + 6, 'rank')
        sht.cell(1, c + 7, 'timeCost')

        objective = objectives[ii]
        rank = np.argsort(results[ii], kind='mergesort')   # 根据平均值进行排序
        for alg in range(len(algorithms)):
            c = 2
            count = 0
            sht.cell(alg + 2, 1, algorithms[alg])
            for t in range(args.num_test_instances):
                sht.cell(alg + 2, c, objective[count][alg])
                c += 1
                count += 1
            sht.cell(alg + 2, c, np.max(objective[:, alg]))
            sht.cell(alg + 2, c + 1, np.min(objective[:, alg]))
            sht.cell(alg + 2, c + 2, np.average(objective[:, alg]))
            sht.cell(alg + 2, c + 3, np.std(objective[:, alg]))
            sht.cell(alg + 2, c + 4, rank[alg])
            sht.cell(alg + 2, c + 5, np.average(test_time[:, alg]))
    xls.save(now_file_name)
